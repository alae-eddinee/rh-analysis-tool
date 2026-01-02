import pandas as pd
import os
import re
import warnings
from datetime import datetime
from openpyxl import load_workbook
import xlrd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# Suppress warnings from openpyxl if it reads misnamed files
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
CHEMIN_DOSSIER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data")
GRAPHIQUE_SORTIE = "Retards_Apres_10AM.png"

# LISTE DES EMPLOYÉS À EXCLURE PAR NOM (Insensible à la casse)
EMPLOYES_EXCLUS = [
    "HMOURI ALI"
]

# CODES QUI SIGNIFIENT UN "OUVRIER"
CODES_OUVRIER = ['130', '140', '141', '131']

# --- CLASSE UTILITAIRE POUR COMPATIBILITÉ XLS ---
class MockCell:
    """Imite un objet cellule openpyxl pour les fichiers .xls lus via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Normalise les noms pour assurer la correspondance malgré les espaces/caractères cachés."""
    if not name:
        return ""
    name = str(name).upper()
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

def extract_month_year_from_filename(file_path):
    """Extrait le mois et l'année du nom de fichier."""
    filename = os.path.basename(file_path).upper()
    
    # Chercher les mois en français dans le nom de fichier
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    # Chercher l'année (4 chiffres)
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    # Chercher le mois
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    # Si aucun mois trouvé, essayer de chercher des nombres de 1-12
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    # Valeur par défaut
    return '12', year

def parse_scan_times(scan_str):
    """Analyse la chaîne pour trouver toutes les entrées de temps (HH:MM)."""
    if scan_str is None:
        return {}, 0, []
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    count = len(times)
    scans = {}
    for i, time_val in enumerate(times):
        scans[f'scan_{i+1}'] = time_val  
    return scans, count, times

def get_sheet_rows(file_path):
    """Générateur qui produit des lignes de fichiers .xlsx ou .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Attention : '{os.path.basename(file_path)}' est un fichier .xlsx nommé comme .xls. Changement de moteur...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Échec de lecture du fichier avec secours : {e2}")
            else:
                print(f"Erreur lors du traitement du fichier .xls {os.path.basename(file_path)} : {e}")
                return

def process_employee_buffer(employee_data):
    """
    Décide si un employé est un OUVRIER basé sur les codes HJ.
    Retourne les enregistrements si ADM, retourne une liste vide si OUVRIER.
    """
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    
    # 1. Filtrer uniquement pour les jours de semaine (exclure Sam/Dim) pour la logique de classification
    weekday_recs = []
    for r in records:
        day_str = str(r.get('day_str', '')).lower()
        if not day_str.startswith('sa') and not day_str.startswith('di'):
            weekday_recs.append(r)
    
    if not weekday_recs:
        return records

    # 2. Compter combien d'enregistrements de jours de semaine correspondent aux CODES_OUVRIER
    ouvrier_matches = 0
    for r in weekday_recs:
        raw_hj = str(r['hj_code'])
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
            
        if hj in CODES_OUVRIER:
            ouvrier_matches += 1
    
    # 3. Calculer le Ratio
    ratio = ouvrier_matches / len(weekday_recs)
    
    # 4. Si strictement plus de 50% correspondent aux codes Ouvrier, les exclure
    if ratio > 0.5:
        return []
    
    return records

def extract_daily_data(file_path):
    """Extrait les données, met en mémoire tampon par employé pour vérifier le statut "Ouvrier" via la colonne HJ."""
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    source_file_name = os.path.basename(file_path)
    month_num, year_num = extract_month_year_from_filename(file_path)
    days_french = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''
            
            # --- VÉRIFIER NOUVELLE SECTION OU NOM (DÉCLENCHE TRAITEMENT TAMpon) ---
            if 'SERVICE / SECTION :' in val_0 or 'NOM :' in val_0:
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 'matricule': '', 'records': []
                }
            
            elif 'NOM :' in val_0:
                raw_name = val_0.replace('NOM :', '').strip()
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name),
                    'matricule': '', 
                    'records': []
                }

            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
                
            # --- ANALYSER LES DONNÉES QUOTIDIENNES ---
            elif any(val_0.startswith(day) for day in days_french) and any(char.isdigit() for char in val_0):
                
                hj_val = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                
                # VÉRIFIER CONGÉ/ABSENCE
                row_text = (val_0 + " " + str(raw_scan_val)).upper()
                if "CONGE-" in row_text:
                    continue

                if 'Date' not in val_0 and 'Heures' not in val_0:
                    scan_times_dict, calculated_count, times_list = parse_scan_times(raw_scan_val)
                    # Tenter d'extraire une date complète (JJ/MM/AAAA)
                    date_match = re.search(r'(\d{1,2})[/](\d{1,2})[/](\d{4})', val_0)
                    full_date = None
                    if date_match:
                        try:
                            d, m, y = map(int, date_match.groups())
                            full_date = datetime(y, m, d)
                        except:
                            pass

                    parts = val_0.split()
                    day_match = re.search(r'\d+', val_0)
                    day_num = int(day_match.group()) if day_match else 0
                    day_str = parts[0] if parts else ''

                    record = {
                        'source_file': source_file_name,
                        'name': current_employee.get('name', ''),
                        'day_raw': parts[0] if parts else '',
                        'day_numeric': day_num,
                        'day_str': day_str,
                        'hj_code': str(hj_val).strip(),
                        'scan_count': calculated_count,
                        'raw_pointages': str(raw_scan_val) if raw_scan_val else '',
                        'month_num': month_num,
                        'year_num': year_num,
                        'date': full_date
                    }
                    current_employee['records'].append(record)
        
        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)

    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier {os.path.basename(file_path)} : {e}")
        return []
    
    return all_records

def is_late_after_10(raw_pointages):
    """Vérifie si le premier scan est après 10:00 AM."""
    scans = re.findall(r'\d{1,2}:\d{2}', str(raw_pointages))
    
    if not scans:
        return False
    
    try:
        first_scan_dt = datetime.strptime(scans[0], '%H:%M')
        limit_1000 = first_scan_dt.replace(hour=10, minute=0, second=0)
        return first_scan_dt > limit_1000
    except:
        return False

def generate_lateness_graph(input_dir, output_dir):
    """
    Génère le graphique des retards à partir des fichiers dans input_dir et le sauvegarde dans output_dir.
    Retourne le chemin du fichier image généré ou None.
    """
    if not os.path.exists(input_dir):
        print(f"Dossier non trouvé : {input_dir}")
        return None
    
    # S'assurer que le dossier de sortie existe
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    all_data = []

    # --- LIRE LES DONNÉES ---
    print("Analyse des fichiers...")
    for file in os.listdir(input_dir):
        # Ignorer les fichiers temporaires Streamlit ou autres
        if file.lower().endswith(('.xls', '.xlsx')) and not file.startswith("Daily_Analysis") and not file.startswith("Monthly") and not file.startswith("Master") and not file.startswith("~$"):
            print(f"Lecture : {file}...")
            full_path = os.path.join(input_dir, file)
            records = extract_daily_data(full_path)
            all_data.extend(records)

    if not all_data:
        print("Aucune donnée valide trouvée.")
        return None

    df = pd.DataFrame(all_data)

    # --- EXCLURE LES EMPLOYÉS PAR NOM ---
    if EMPLOYES_EXCLUS:
        print(f"\nFiltrage des noms exclus : {EMPLOYES_EXCLUS}")
        excluded_clean = [clean_name_string(name) for name in EMPLOYES_EXCLUS]
        initial_len = len(df)
        df = df[~df['name'].isin(excluded_clean)]
        print(f"Supprimé {initial_len - len(df)} enregistrements basés sur la liste d'exclusion de noms.")
    
    if df.empty:
        print("Toutes les données filtrées.")
        return None

    # --- ATTRIBUTION DES DATES ---
    # Si des dates ont été extraites directement, on les utilise.
    # Sinon (anciens formats), on applique la logique de pivot par fichier.
    files_to_process = df['source_file'].unique()
    final_records = []

    for f in files_to_process:
        df_file = df[df['source_file'] == f].copy()
        
        # Vérifier si on a déjà des dates valides pour ce fichier
        if df_file['date'].notnull().any():
            final_records.append(df_file)
            continue
            
        # Sinon, logique de fallback (pivot)
        month_num = df_file['month_num'].iloc[0]
        year_num = df_file['year_num'].iloc[0]
        
        unique_days = []
        seen = set()
        for d in df_file['day_numeric']:
            if d not in seen:
                unique_days.append(d)
                seen.add(d)
        
        has_transition = False
        pivot_index = -1
        for i in range(len(unique_days) - 1):
            if unique_days[i] > unique_days[i+1]:
                has_transition = True
                pivot_index = i
                break
        
        base_month = int(month_num)
        base_year = int(year_num)
        day_to_date = {}
        
        for i, d in enumerate(unique_days):
            try:
                if has_transition:
                    if i <= pivot_index:
                        m = base_month - 1
                        y = base_year
                        if m == 0: m = 12; y -= 1
                        day_to_date[d] = datetime(y, m, d)
                    else:
                        day_to_date[d] = datetime(base_year, base_month, d)
                else:
                    day_to_date[d] = datetime(base_year, base_month, d)
            except:
                continue
        
        df_file['date'] = df_file['day_numeric'].map(day_to_date)
        final_records.append(df_file)

    df = pd.concat(final_records)
    df = df.dropna(subset=['date'])

    if df.empty:
        print("Aucune donnée valide après traitement des dates.")
        return None

    # --- ÉLAGUER LE DERNIER JOUR SI INCOMPLET ---
    unique_dates = sorted(df['date'].unique())
    if unique_dates:
        last_date = unique_dates[-1]
        last_day_records = df[df['date'] == last_date]
        total = len(last_day_records)
        incomplete = len(last_day_records[last_day_records['scan_count'] <= 1])
        
        if total > 0 and (incomplete / total) > 0.5 and len(unique_dates) > 1:
            print(f"DÉCISION : Le jour {last_date.strftime('%d/%m')} est incomplet. Supprimé.")
            df = df[df['date'] != last_date].copy()
            unique_dates = unique_dates[:-1]

    if df.empty:
        print("Aucune donnée valide trouvée après filtrage des dates.")
        return None

    # --- TROUVER LA PLAGE DE DATES DISPONIBLE ---
    min_date = df['date'].min()
    max_date = df['date'].max()
    print(f"Période finale : du {min_date.strftime('%d/%m/%Y')} au {max_date.strftime('%d/%m/%Y')}")

    # --- CALCULER LES RETARDS ---
    print("\nCalcul des retards après 10:00 AM...")
    df['is_late_1000'] = df['raw_pointages'].apply(is_late_after_10)
    
    # Grouper par date et compter les retards
    daily_late_count = df[df['is_late_1000']].groupby('date').size().reset_index(name='late_count')
    
    # --- COMPLÉTER LES JOURS MANQUANTS (Sundays, holidays) ---
    all_dates = pd.date_range(start=df['date'].min(), end=df['date'].max())
    all_days = pd.DataFrame({'date': all_dates})
    
    daily_late_count = all_days.merge(daily_late_count, on='date', how='left').fillna(0)
    daily_late_count = daily_late_count.sort_values('date')
    
    # --- CRÉER LE GRAPHIQUE ---
    print("\nGénération du graphique...")
    plt.figure(figsize=(14, 7))
    
    # Créer un graphique en barres
    plt.bar(daily_late_count['date'], daily_late_count['late_count'], 
            color='#ED7D31', edgecolor='black', linewidth=0.5, alpha=0.8)
    
    # Ajouter un graphique linéaire pour la tendance
    plt.plot(daily_late_count['date'], daily_late_count['late_count'], 
             color='#C00000', marker='o', linewidth=2, markersize=6, label='Tendance')
    
    # Formatage du titre avec la période exacte
    start_str = min_date.strftime('%d %b %Y')
    end_str = max_date.strftime('%d %b %Y')
    
    if min_date.year == max_date.year and min_date.month == max_date.month:
        # Même mois
        months_fr = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
                     'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        month_name = months_fr[min_date.month - 1]
        period_title = f"{month_name} {min_date.year}"
    else:
        # Période couvrant plusieurs mois
        period_title = f"Du {start_str} au {end_str}"

    plt.title(f"Nombre d'Employés Arrivant Après 10:00 AM\n{period_title}", 
              fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Date', fontsize=12, fontweight='bold')
    plt.ylabel('Nombre de Retards (Après 10:00)', fontsize=12, fontweight='bold')
    
    # Formater l'axe des x pour afficher chaque date individuellement
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%d %b'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator(interval=1)) 
    plt.xticks(rotation=45, ha='right', fontsize=9)
    
    # Ajouter une grille pour une meilleure lisibilité
    plt.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Ajouter des étiquettes de valeur au-dessus des barres
    for idx, row in daily_late_count.iterrows():
        if row['late_count'] > 0:
            plt.text(row['date'], row['late_count'] + 0.3, 
                    f"{int(row['late_count'])}", 
                    ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    plt.legend()
    plt.tight_layout()
    
    # Sauvegarder le graphique
    output_path = os.path.join(output_dir, GRAPHIQUE_SORTIE)
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"\nSUCCÈS ! Graphique sauvegardé : {output_path}")
    
    # Afficher les statistiques
    print("\n--- STATISTIQUES ---")
    print(f"Total des jours analysés : {len(daily_late_count)}")
    print(f"Total des retards (après 10:00) : {int(daily_late_count['late_count'].sum())}")
    print(f"Moyenne des retards par jour : {daily_late_count['late_count'].mean():.2f}")
    if not daily_late_count.empty:
        print(f"Maximum de retards en un jour : {int(daily_late_count['late_count'].max())}")
        idx_max = daily_late_count['late_count'].idxmax()
        worst_day = daily_late_count.loc[idx_max, 'date']
        print(f"Jour avec le plus de retards : {worst_day.strftime('%d %B %Y')}")
    
    plt.close() # Fermer la figure pour libérer la mémoire
    
    return output_path

def main():
    if not os.path.exists(CHEMIN_DOSSIER):
        print(f"Dossier non trouvé : {CHEMIN_DOSSIER}")
        return
    
    # Mode standalone
    output = generate_lateness_graph(CHEMIN_DOSSIER, CHEMIN_DOSSIER)
    if output:
        print(f"Image générée : {output}")

if __name__ == "__main__":
    main()
