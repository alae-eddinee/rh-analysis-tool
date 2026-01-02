import pandas as pd
import os
import re
import warnings
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd

# Supprimer les avertissements de openpyxl si il lit des fichiers mal nommés
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
CHEMIN_DOSSIER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data")
NOM_FICHIER_SORTIE = "Analyse_Quotidienne_Rapport_Avec_Comptages.xlsx"

# LISTE DES EMPLOYÉS À EXCLURE PAR NOM (Insensible à la casse)
EMPLOYES_EXCLUS = [
    # "ABOU HASNAA", 
    "HMOURI ALI"
]

# CODES QUI SIGNIFIENT UN "OUVRIER"
CODES_OUVRIER = ['130', '140', '141', '131']

# --- CLASSE UTILITAIRE POUR COMPATIBILITÉ XLS ---
class MockCell:
    """Imite un objet cellule openpyxl pour les fichiers .xls lus via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Normalise les noms pour assurer la correspondance malgré les espaces/caractères cachés."""
    if not name:
        return ""
    name = str(name).upper()
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

def parse_scan_times(scan_str):
    """Analyse la chaîne pour trouver toutes les entrées de temps (HH:MM)."""
    if scan_str is None:
        return {}, 0, []
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    count = len(times)
    scans = {}
    for i, time_val in enumerate(times):
        scans[f'scan_{i+1}'] = time_val  
    return scans, count, times

def get_sheet_rows(file_path):
    """Générateur qui produit des lignes de fichiers .xlsx ou .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Attention : '{os.path.basename(file_path)}' est un fichier .xlsx nommé comme .xls. Changement de moteur...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Échec de lecture du fichier avec secours : {e2}")
            else:
                print(f"Erreur lors du traitement du fichier .xls {os.path.basename(file_path)} : {e}")
                return

def process_employee_buffer(employee_data):
    """
    Décide si un employé est un OUVRIER basé sur les codes HJ.
    Retourne les enregistrements si ADM, retourne une liste vide si OUVRIER.
    """
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    name = employee_data.get('name', 'Unknown')
    
    # 1. Filtrer uniquement pour les jours de semaine (exclure Sam/Dim) pour la logique de classification
    weekday_recs = []
    for r in records:
        day_str = str(r.get('day_str', '')).lower()
        if not day_str.startswith('sa') and not day_str.startswith('di'):
            weekday_recs.append(r)
    
    if not weekday_recs:
        return records

    # 2. Compter combien d'enregistrements de jours de semaine correspondent aux CODES_OUVRIER
    ouvrier_matches = 0
    for r in weekday_recs:
        raw_hj = str(r['hj_code'])
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
            
        if hj in CODES_OUVRIER:
            ouvrier_matches += 1
    
    # 3. Calculer le Ratio
    ratio = ouvrier_matches / len(weekday_recs)
    
    # 4. Si strictement plus de 50% correspondent aux codes Ouvrier, les exclure
    if ratio > 0.5:
        return []
    
    return records

def extract_month_year_from_filename(file_path):
    """Extrait le mois et l'année du nom de fichier."""
    filename = os.path.basename(file_path).upper()
    
    # Chercher les mois en français dans le nom de fichier
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    # Chercher l'année (4 chiffres)
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    # Chercher le mois
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    # Si aucun mois trouvé, essayer de chercher des nombres de 1-12
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    # Valeur par défaut
    return '12', year

def extract_daily_data(file_path):
    """Extrait les données, met en mémoire tampon par employé pour vérifier le statut "Ouvrier" via la colonne HJ."""
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    source_file_name = os.path.basename(file_path)
    month_num, year_num = extract_month_year_from_filename(file_path)
    days_french = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''
            
            # --- VÉRIFIER NOUVELLE SECTION OU NOM (DÉCLENCHE TRAITEMENT TAMpon) ---
            if 'SERVICE / SECTION :' in val_0 or 'NOM :' in val_0:
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 'matricule': '', 'records': []
                }
            
            elif 'NOM :' in val_0:
                raw_name = val_0.replace('NOM :', '').strip()
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name),
                    'matricule': '', 
                    'records': []
                }

            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
                
            # --- ANALYSER LES DONNÉES QUOTIDIENNES ---
            elif any(val_0.startswith(day) for day in days_french) and any(char.isdigit() for char in val_0):
                
                hj_val = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                
                # VÉRIFIER CONGÉ/ABSENCE
                row_text = (val_0 + " " + str(raw_scan_val)).upper()
                if "CONGE-" in row_text:
                    continue

                if 'Date' not in val_0 and 'Heures' not in val_0:
                    scan_times_dict, calculated_count, times_list = parse_scan_times(raw_scan_val)
                    parts = val_0.split()
                    
                    day_match = re.search(r'\d+', val_0)
                    day_num = int(day_match.group()) if day_match else 0
                    day_str = parts[0] if parts else ''

                    record = {
                        'source_file': source_file_name,
                        'name': current_employee.get('name', ''),
                        'day_raw': parts[0] if parts else '',
                        'day_numeric': day_num,
                        'day_str': day_str,
                        'hj_code': str(hj_val).strip(),
                        'scan_count': calculated_count,
                        'raw_pointages': str(raw_scan_val) if raw_scan_val else '',
                        'month_num': month_num,
                        'year_num': year_num
                    }
                    current_employee['records'].append(record)
        
        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)

    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier {os.path.basename(file_path)} : {e}")
        return []
    
    return all_records

def analyze_row(row):
    """Calcule les indicateurs pour retard, pas de déjeuner, heures et demi-journée."""
    scans = re.findall(r'\d{1,2}:\d{2}', str(row.get('raw_pointages', '')))
    
    # Initialiser les valeurs par défaut
    late_930 = False
    late_1000 = False
    late_1400 = False
    no_lunch = False
    is_half_day = False
    hours_worked = 0.0

    if not scans:
        return late_930, late_1000, late_1400, no_lunch, hours_worked, is_half_day
    
    # --- CALCUL DE DURÉE ---
    total_seconds = 0
    for i in range(0, len(scans) - 1, 2):
        t_in = datetime.strptime(scans[i], '%H:%M')
        t_out = datetime.strptime(scans[i+1], '%H:%M')
        if t_out < t_in: t_out += timedelta(days=1)
        total_seconds += (t_out - t_in).total_seconds()
        
    hours_worked = round(total_seconds / 3600, 2)
    
    # --- LOGIQUE DE RETARD (Hiérarchie Stricte) ---
    first_scan_dt = datetime.strptime(scans[0], '%H:%M')
    
    limit_930 = first_scan_dt.replace(hour=9, minute=30, second=0)
    limit_1000 = first_scan_dt.replace(hour=10, minute=0, second=0)
    limit_1400 = first_scan_dt.replace(hour=14, minute=0, second=0)

    # Priorité 1 : Retard après 14:00 (Prend le dessus)
    if first_scan_dt > limit_1400:
        late_1400 = True
        late_1000 = False # Assurer aucune duplication
        late_930 = False
    # Priorité 2 : Retard après 10:00 (mais avant 14:00)
    elif first_scan_dt > limit_1000:
        late_1400 = False
        late_1000 = True
        late_930 = False
    # Priorité 3 : Retard après 09:30 (mais avant 10:00)
    elif first_scan_dt > limit_930:
        late_1400 = False
        late_1000 = False
        late_930 = True

    # --- VÉRIFICATION PAS DE DÉJEUNER ---
    # Si début d'après-midi, Pas de Déjeuner n'est pas applicable/déjà signalé par Retard 14h
    if late_1400:
        no_lunch = False
    else:
        no_lunch = len(scans) < 4 and len(scans) > 0
    
    # --- LOGIQUE DEMI-JOURNÉE ---
    # Règles : 
    # 1. Pas Samedi.
    # 2. Entrée >= 13:00 (Après-midi Seulement) OU (Sortie <= 14:00 ET Heures < 7) (Matin Seulement)
    
    day_str = str(row.get('day_str', '')).lower()
    is_saturday = day_str.startswith('sa')

    if not is_saturday and len(scans) >= 2 and hours_worked > 0:
        t_first = datetime.strptime(scans[0], '%H:%M')
        t_last = datetime.strptime(scans[-1], '%H:%M')
        
        # Gérer la logique de quart de nuit au cas où
        if t_last < t_first: 
            t_last += timedelta(days=1)
            
        limit_1300 = t_first.replace(hour=13, minute=0, second=0)
        # Redéfinir limit_1400 basé sur t_first pour la cohérence du type d'objet, bien que les heures comptent le plus
        limit_1400_exit = t_first.replace(hour=14, minute=0, second=0)
        
        # Condition A : Entré >= 13:00 (Quart d'après-midi / Retard)
        cond_afternoon = (t_first >= limit_1300)
        
        # Condition B : Parti <= 14:00 (Quart de matin / Départ anticipé) ET Heures < 7
        cond_morning = (t_last <= limit_1400_exit) and (hours_worked < 7.0)
        
        if cond_afternoon or cond_morning:
            is_half_day = True
    
    return late_930, late_1000, late_1400, no_lunch, hours_worked, is_half_day

def create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, flag_column, output_header):
    """Crée un DataFrame à 3 colonnes : [Nom, Compte, %]"""
    subset = daily_df[daily_df[flag_column]].copy()
    result = subset[['name']].reset_index(drop=True)
    
    # Utiliser les bonnes statistiques selon le type de jour cible
    if flag_column == 'is_under_hours' and daily_df['day_str'].iloc[0].startswith('Sa'):
        # Pour les samedis, utiliser les statistiques de samedi
        stats_to_use = monthly_stats_saturday
    else:
        # Pour les autres cas, utiliser les statistiques combinées
        stats_to_use = monthly_stats
    
    # Mapper les comptes depuis les statistiques mensuelles appropriées
    result['Count'] = result['name'].map(stats_to_use[flag_column]).fillna(0).astype(int)
    
    # Total des jours basé sur HEURES > 0 (Logique stricte "Jour Travailé" correspondant au script mensuel)
    total_days = result['name'].map(stats_to_use['total_attendance']).fillna(1) 
    
    result['%'] = (result['Count'] / total_days)
    
    result.columns = [output_header, 'Count', '%']
    return result

def process_daily_analysis(input_dir, output_dir):
    """
    Traite les fichiers dans input_dir et sauvegarde l'analyse dans output_dir.
    Retourne le chemin du fichier généré ou None.
    """
    if not os.path.exists(input_dir):
        print(f"Dossier non trouvé : {input_dir}")
        return None
    
    # S'assurer que le dossier de sortie existe
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_data = []

    # --- LIRE LES DONNÉES ---
    print("Analyse des fichiers...")
    for file in os.listdir(input_dir):
        # Ignorer les fichiers temporaires Streamlit ou autres
        if file.lower().endswith(('.xls', '.xlsx')) and not file.startswith("Daily_Analysis") and not file.startswith("Monthly") and not file.startswith("Master") and not file.startswith("~$"):
            print(f"Lecture : {file}...")
            full_path = os.path.join(input_dir, file)
            records = extract_daily_data(full_path)
            all_data.extend(records)

    if not all_data:
        print("Aucune donnée valide trouvée.")
        return None

    df = pd.DataFrame(all_data)

    # --- EXCLURE LES EMPLOYÉS PAR NOM ---
    if EMPLOYES_EXCLUS:
        print(f"\nFiltrage des noms exclus : {EMPLOYES_EXCLUS}")
        excluded_clean = [clean_name_string(name) for name in EMPLOYES_EXCLUS]
        initial_len = len(df)
        df = df[~df['name'].isin(excluded_clean)]
        print(f"Supprimé {initial_len - len(df)} enregistrements basés sur la liste d'exclusion de noms.")
    
    if df.empty:
        print("Toutes les données filtrées.")
        return None

    # --- DÉTECTION CHRONOLOGIQUE AMÉLIORÉE ---
    if 'day_numeric' in df.columns and not df.empty:
        # 1. Récupérer les infos de base
        month_num = df['month_num'].iloc[0] if 'month_num' in df.columns else '01'
        year_num = df['year_num'].iloc[0] if 'year_num' in df.columns else '2026'
        
        # 2. Identifier la séquence chronologique réelle
        unique_days_in_order = []
        seen = set()
        for d in df['day_numeric']:
            if d not in seen:
                unique_days_in_order.append(d)
                seen.add(d)

        real_start_day = unique_days_in_order[0]
        real_end_day = unique_days_in_order[-1]
        
        # Détecter s'il y a une transition de mois (ex: 25, 26... 31, 1, 2)
        has_transition = False
        pivot_index = -1
        for i in range(len(unique_days_in_order) - 1):
            if unique_days_in_order[i] > unique_days_in_order[i+1]:
                has_transition = True
                pivot_index = i
                break
        
        print(f"\n--- ANALYSE DE LA PÉRIODE ---")
        print(f"Séquence détectée : {unique_days_in_order}")
        
        # 3. Définir le jour cible (le dernier jour chronologique)
        target_report_day = real_end_day
        
        # 4. Vérifier si le dernier jour est complet (Scan count)
        last_day_records = df[df['day_numeric'] == target_report_day]
        total_last_day = len(last_day_records)
        # On considère un jour incomplet si + de 50% des gens n'ont qu'un seul pointage (ou 0)
        incomplete_count = len(last_day_records[last_day_records['scan_count'] <= 1])
        
        if total_last_day > 0 and (incomplete_count / total_last_day) > 0.5:
            print(f"DÉCISION : Le jour {target_report_day} est incomplet (en cours).")
            # Supprimer le jour incomplet du DataFrame pour l'analyse
            df = df[df['day_numeric'] != target_report_day].copy()
            # Le nouveau jour cible devient le précédent dans la liste ordonnée
            if len(unique_days_in_order) > 1:
                target_report_day = unique_days_in_order[-2]
                real_end_day = target_report_day
            print(f"Nouveau jour cible : {target_report_day}")
        else:
            print(f"DÉCISION : Le jour {target_report_day} est complet.")

        # 5. Calcul du nom du mois pour le header
        month_names = {
            '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
            '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
            '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
        }
        month_name = month_names.get(month_num, f'Mois {month_num}')
        
    else:
        print("Erreur: Aucune donnée numérique de jour trouvée.")
        return None

    # --- CALCUL DES MÉTRIQUES ---
    print("\nCalcul des métriques...")
    results = df.apply(analyze_row, axis=1)
    
    df['is_late_930'] = [x[0] for x in results]
    df['is_late_1000'] = [x[1] for x in results]
    df['is_late_1400'] = [x[2] for x in results]
    df['no_lunch'] = [x[3] for x in results]
    df['hours_worked'] = [x[4] for x in results]
    df['is_half_day'] = [x[5] for x in results] 
    
    if 'day_str' in df.columns:
        mask_saturday = df['day_str'].astype(str).str.startswith('Sa')
        df.loc[mask_saturday, 'no_lunch'] = False

    df['target_hours'] = df['day_str'].apply(lambda x: 4.0 if str(x).startswith('Sa') else 8.0)
    df['is_under_hours'] = (df['scan_count'] > 0) & (df['hours_worked'] < df['target_hours'])

    # --- GÉNÉRATION DES STATISTIQUES ---
    cols_to_sum = ['is_late_930', 'is_late_1000', 'is_late_1400', 'no_lunch', 'is_half_day']
    
    valid_days_df = df[df['hours_worked'] > 0]
    
    saturday_records = valid_days_df[valid_days_df['day_str'].str.startswith('Sa')]
    weekday_records = valid_days_df[~valid_days_df['day_str'].str.startswith('Sa')]
    
    monthly_stats_weekday = weekday_records.groupby('name')[cols_to_sum].sum()
    monthly_stats_weekday['total_attendance'] = weekday_records.groupby('name').size()
    monthly_stats_weekday['is_under_hours'] = weekday_records.groupby('name')['is_under_hours'].sum()
    
    monthly_stats_saturday = saturday_records.groupby('name')[cols_to_sum].sum()
    monthly_stats_saturday['total_attendance'] = saturday_records.groupby('name').size()
    monthly_stats_saturday['is_under_hours'] = saturday_records.groupby('name')['is_under_hours'].sum()
    
    monthly_stats = monthly_stats_weekday.combine_first(monthly_stats_saturday)
    
    # --- FILTRER POUR LE JOUR CIBLE DU RAPPORT ---
    if 'day_numeric' in df.columns:
        daily_df = df[df['day_numeric'] == target_report_day].copy()
    else:
        daily_df = pd.DataFrame()

    if daily_df.empty:
        print(f"\nATTENTION : Aucun enregistrement trouvé pour le Jour {target_report_day}.")
        return None

    sample_day_str = daily_df.iloc[0]['day_str'] if not daily_df.empty else ""
    is_target_saturday = str(sample_day_str).startswith('Sa')

    # --- PRÉPARER LES LISTES DE SORTIE ---
    under_header = "Moins de 4h" if is_target_saturday else "Moins de 8h"
    df_under = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_under_hours', under_header)

    df_late_10 = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_late_1000', "Entrée > 10:00")
    df_late_930 = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_late_930', "Entrée > 09:30")
    df_late_1400 = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_late_1400', "Entrée > 14:00")
    df_half_day = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'is_half_day', "Demi-Journée")

    if is_target_saturday:
        main_list = pd.concat([df_under, df_late_10, df_late_930, df_late_1400], axis=1)
    else:
        df_no_lunch = create_category_dataframe(daily_df, monthly_stats, monthly_stats_saturday, 'no_lunch', "Pas de Déjeuner")
        main_list = pd.concat([df_under, df_half_day, df_no_lunch, df_late_10, df_late_930, df_late_1400], axis=1)

    # --- EXPORTER VERS EXCEL ---
    # Calculer la plage de jours analysés
    if not df.empty and 'day_numeric' in df.columns:
        if has_transition:
            first_month_days = unique_days_in_order[:pivot_index + 1]
            second_month_days = unique_days_in_order[pivot_index + 1:]
            total_days = len(first_month_days) + len(second_month_days)
        else:
            total_days = len(unique_days_in_order)
        
        # Créer un nom de fichier dynamique basé sur la période analysée
        dynamic_filename = f"POINTAGE ANALYSE DU {real_start_day:02d}-{month_num}-{year_num} A {real_end_day:02d}-{month_num}-{year_num}.xlsx"
        output_path = os.path.join(output_dir, dynamic_filename)
        header_text = f"Analyse Quotidienne - Période : {real_start_day} au {real_end_day} {month_name} {year_num}"
    else:
        header_text = "Analyse Quotidienne - Période non spécifiée"
        output_path = os.path.join(output_dir, NOM_FICHIER_SORTIE)
    
    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            # Ajouter l'en-tête sur la première ligne
            main_list.to_excel(writer, sheet_name='Analyse Quotidienne', index=False, header=False, startrow=2)
            
            workbook = writer.book
            worksheet = writer.sheets['Analyse Quotidienne']
            
            # Format pour l'en-tête de période
            header_title = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 14, 'font_color': '#2F5597', 'border': 1
            })
            
            # Écrire l'en-tête de période sur la première ligne (fusionnée)
            if len(main_list.columns) > 1:
                worksheet.merge_range(0, 0, 0, len(main_list.columns) - 1, header_text, header_title)
            else:
                worksheet.write(0, 0, header_text, header_title)
            
            # Formats
            header_blue = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            header_orange = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'fg_color': '#ED7D31', 'font_color': 'white', 'border': 1
            })
            header_red = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'fg_color': '#C00000', 'font_color': 'white', 'border': 1
            })
            
            body_left = workbook.add_format({'border': 1, 'align': 'left'})
            body_center = workbook.add_format({'border': 1, 'align': 'center'})
            body_pct = workbook.add_format({'border': 1, 'align': 'center', 'num_format': '0%'})

            max_rows = len(main_list)
            columns = main_list.columns.tolist()

            for i, col_name in enumerate(columns):
                col_name_str = str(col_name)
                
                # Formatage Dynamique d'En-tête
                col_format = body_left
                header_style = header_blue

                if "Count" in col_name_str:
                    header_style = header_orange
                    col_format = body_center
                elif "%" in col_name_str:
                    header_style = header_orange
                    col_format = body_pct
                elif "14:00" in col_name_str:
                    header_style = header_red
                elif "Demi-Journée" in col_name_str:
                    header_style = header_orange
                
                worksheet.write(1, i, col_name, header_style)

                col_data = main_list.iloc[:, i]
                max_data_len = 0
                if "%" in col_name_str:
                    max_data_len = 5
                else:
                    valid_data = col_data.dropna().astype(str)
                    if not valid_data.empty:
                        max_data_len = valid_data.map(len).max()
                
                final_width = max(max_data_len, len(col_name_str)) + 4
                worksheet.set_column(i, i, final_width)

                for row_idx in range(max_rows):
                    cell_val = main_list.iloc[row_idx, i]
                    if pd.isna(cell_val):
                        worksheet.write(row_idx + 2, i, "", col_format)
                    else:
                        worksheet.write(row_idx + 2, i, cell_val, col_format)

        print(f"\nSUCCÈS ! Rapport sauvegardé : {output_path}")
        return output_path

    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier : {e}")
        return None

def main():
    if not os.path.exists(CHEMIN_DOSSIER):
        print(f"Dossier non trouvé : {CHEMIN_DOSSIER}")
        return
    
    # Mode standalone (utilisation classique)
    output = process_daily_analysis(CHEMIN_DOSSIER, CHEMIN_DOSSIER)
    if output:
        print(f"Fichier généré : {output}")

if __name__ == "__main__":
    main()