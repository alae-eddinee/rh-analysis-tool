import pandas as pd
import os
import re
import warnings
from datetime import datetime, timedelta
from openpyxl import load_workbook
import xlrd

# Suppress warnings from openpyxl if it reads misnamed files
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- CONFIGURATION ---
import os
FOLDER_PATH = os.path.join(os.path.dirname(__file__), "Data")
OUTPUT_FILENAME = "Monthly_Global_Analysis.xlsx"

# LIST OF EMPLOYEES TO EXCLUDE (Case insensitive)
EXCLUDED_EMPLOYEES = [
    # "ABOU HASNAA", 
    "HMOURI ALI"
]

# CODES THAT SIGNIFY AN "OUVRIER" (Worker)
OUVRIER_CODES = ['130', '140', '141', '131']

# Days of week mapping
DAYS_FRENCH = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']

# --- HELPER CLASS FOR XLS COMPATIBILITY ---
class MockCell:
    """Mimics an openpyxl cell object for .xls files read via xlrd."""
    def __init__(self, value):
        self.value = value

def clean_name_string(name):
    """Normalizes names to ensure matching works despite spaces/hidden chars."""
    if not name:
        return ""
    name = str(name).upper()
    name = name.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

def parse_scan_times(scan_str):
    """Parses scan string to count scans and calculate duration."""
    if scan_str is None:
        return [], 0
    scan_str = str(scan_str)
    times = re.findall(r'\d{1,2}:\d{2}', scan_str)
    return times, len(times)

def calculate_hours_from_scans(times):
    """Calculates total worked hours from a list of 'HH:MM' strings."""
    if not times:
        return 0.0
    
    total_seconds = 0
    for i in range(0, len(times) - 1, 2):
        try:
            t_in = datetime.strptime(times[i], '%H:%M')
            t_out = datetime.strptime(times[i+1], '%H:%M')
            if t_out < t_in: t_out += timedelta(days=1)
            total_seconds += (t_out - t_in).total_seconds()
        except:
            continue
            
    return round(total_seconds / 3600, 2)

def calculate_lunch_minutes(times):
    """Calculates the duration of the lunch break (gap between scan 2 and 3)."""
    if not times or len(times) < 4:
        return 0
    
    try:
        t_out_lunch = datetime.strptime(times[1], '%H:%M')
        t_in_lunch = datetime.strptime(times[2], '%H:%M')
        
        if t_in_lunch < t_out_lunch: 
            t_in_lunch += timedelta(days=1)
            
        diff_seconds = (t_in_lunch - t_out_lunch).total_seconds()
        return diff_seconds / 60 
    except:
        return 0

def get_sheet_rows(file_path):
    """Generator that yields rows from either .xlsx or .xls."""
    ext = os.path.splitext(file_path)[1].lower()
    
    def read_with_openpyxl(path):
        wb = load_workbook(path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows():
            yield row

    if ext in ['.xlsx', '.xlsm']:
        yield from read_with_openpyxl(file_path)
    elif ext == '.xls':
        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    val = sheet.cell_value(row_idx, col_idx)
                    row_data.append(MockCell(val))
                yield row_data
        except Exception as e:
            error_msg = str(e).lower()
            if "xlsx" in error_msg or "zip" in error_msg:
                print(f"Warning: '{os.path.basename(file_path)}' is an .xlsx file named as .xls. Switching engine...")
                try:
                    yield from read_with_openpyxl(file_path)
                except Exception as e2:
                    print(f"Failed to read file with fallback: {e2}")
            else:
                print(f"Error processing .xls file {os.path.basename(file_path)}: {e}")
                return

def process_employee_buffer(employee_data):
    """Decides if an employee is an OUVRIER based on HJ codes."""
    if not employee_data or not employee_data.get('records'):
        return []

    records = employee_data['records']
    weekday_recs = [r for r in records if not str(r['day_str']).startswith(('Sa', 'Di'))]
    
    if not weekday_recs:
        return records

    ouvrier_matches = 0
    for r in weekday_recs:
        raw_hj = str(r.get('hj_code', ''))
        if '.' in raw_hj:
            hj = raw_hj.split('.')[0].strip()
        else:
            hj = raw_hj.strip()
            
        if hj in OUVRIER_CODES:
            ouvrier_matches += 1
    
    ratio = ouvrier_matches / len(weekday_recs)
    if ratio > 0.5:
        return []
    
    return records

def extract_month_year_from_filename(file_path):
    """Extracts month and year from filename."""
    filename = os.path.basename(file_path).upper()
    
    # Look for French months in filename
    months = {
        'JANVIER': '01', 'FEVRIER': '02', 'MARS': '03', 'AVRIL': '04',
        'MAI': '05', 'JUIN': '06', 'JUILLET': '07', 'AOUT': '08',
        'SEPTEMBRE': '09', 'OCTOBRE': '10', 'NOVEMBRE': '11', 'DECEMBRE': '12'
    }
    
    # Look for year (4 digits)
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = year_match.group(1) if year_match else '2025'
    
    # Look for month
    for month_name, month_num in months.items():
        if month_name in filename:
            return month_num, year
    
    # If no month found, try to find numbers 1-12
    month_match = re.search(r'\b(0[1-9]|1[0-2])\b', filename)
    if month_match:
        return month_match.group(1), year
    
    # Default value
    return '12', year

def extract_date_from_string(date_str):
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(date_str))
    if match:
        try:
            return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except:
            return None
    return None

def extract_data(file_path):
    all_records = []
    current_employee = {'service': '', 'name': '', 'matricule': '', 'records': []}
    month_num, year_num = extract_month_year_from_filename(file_path)
    
    try:
        for row in get_sheet_rows(file_path):
            if not row: continue
            
            cell_0 = row[0]
            val_0 = str(cell_0.value).strip() if cell_0.value else ''

            if 'SERVICE / SECTION :' in val_0 or 'NOM :' in val_0:
                valid_records = process_employee_buffer(current_employee)
                all_records.extend(valid_records)

            if 'SERVICE / SECTION :' in val_0:
                current_employee = {
                    'service': val_0.replace('SERVICE / SECTION :', '').strip(),
                    'name': '', 
                    'matricule': '',
                    'records': []
                }
            elif 'NOM :' in val_0:
                raw_name = val_0.replace('NOM :', '').strip()
                current_employee = {
                    'service': current_employee.get('service', ''),
                    'name': clean_name_string(raw_name), 
                    'matricule': '',
                    'records': []
                }
            elif 'MATRICULE :' in val_0:
                current_employee['matricule'] = val_0.replace('MATRICULE :', '').strip()
            elif any(val_0.startswith(day) for day in DAYS_FRENCH) and any(char.isdigit() for char in val_0):
                hj_val = row[1].value if len(row) > 1 else ''
                raw_scan_val = row[2].value if len(row) > 2 else ''
                row_text_upper = (str(val_0) + " " + str(raw_scan_val)).upper()
                date_obj = extract_date_from_string(val_0)
                
                is_leave = 0
                is_holiday = 0
                is_day_worked = 0
                hours_worked = 0.0
                daily_target_for_worked_day = 0.0 
                daily_lunch_minutes = 0
                has_lunch_break = 0 
                times_list = []
                scan_count = 0

                is_saturday = val_0.lower().startswith('sa')
                is_sunday = val_0.lower().startswith('di')

                if "JOUR FERIE" in row_text_upper:
                    is_holiday = 1
                    if is_sunday: is_holiday = 0 
                elif "CONGE" in row_text_upper:
                    is_leave = 1
                elif "ABSENCE NON JUSTIFIÉE-" in row_text_upper:
                    pass 
                else:
                    times_list, scan_count = parse_scan_times(raw_scan_val)
                    hours_worked = calculate_hours_from_scans(times_list)
                    
                    if len(times_list) >= 4 and not is_saturday:
                        daily_lunch_minutes = calculate_lunch_minutes(times_list)
                        has_lunch_break = 1
                    
                    if hours_worked > 0:
                        is_day_worked = 1
                        if is_saturday:
                            daily_target_for_worked_day = 4.0
                        else:
                            daily_target_for_worked_day = 8.0

                if date_obj:
                    day_numeric = date_obj.day
                    day_str = val_0.split()[0] if val_0 else ''

                    record = {
                        'name': current_employee.get('name', ''),
                        'service': current_employee.get('service', ''),
                        'full_date': date_obj, 
                        'day_numeric': day_numeric,
                        'day_str': day_str,
                        'hj_code': str(hj_val).strip(),
                        'times_list': times_list,
                        'hours_worked': hours_worked,
                        'is_day_worked': is_day_worked,
                        'is_leave': is_leave,
                        'is_holiday': is_holiday,
                        'scan_count': scan_count,
                        'daily_target_for_worked_day': daily_target_for_worked_day,
                        'daily_lunch_minutes': daily_lunch_minutes,
                        'has_lunch_break': has_lunch_break,
                        'month_num': month_num,
                        'year_num': year_num
                    }
                    current_employee['records'].append(record)

        valid_records = process_employee_buffer(current_employee)
        all_records.extend(valid_records)
    
    except Exception as e:
        print(f"Error opening {os.path.basename(file_path)}: {e}")
        return []
    
    return all_records

def analyze_record(row):
    """Applies business rules to a single daily record."""
    # Initialize Defaults
    is_late_930 = 0
    is_late_1000 = 0
    is_late_1400 = 0
    no_lunch = 0
    is_under = 0
    is_half_day = 0

    if row['is_leave'] or row['is_holiday']:
        return 0, 0, 0, 0, 0, 0

    times = row['times_list']
    if not times: 
        return 0, 0, 0, 0, 0, 0

    first_scan = datetime.strptime(times[0], '%H:%M')
    
    # --- TIME LIMITS ---
    limit_930 = first_scan.replace(hour=9, minute=30, second=0)
    limit_1000 = first_scan.replace(hour=10, minute=0, second=0)
    limit_1300 = first_scan.replace(hour=13, minute=0, second=0)
    limit_1400 = first_scan.replace(hour=14, minute=0, second=0)

    # --- LATENESS LOGIC (ANTI-DUPLICATION) ---
    if first_scan > limit_1400:
        is_late_1400 = 1
        is_late_1000 = 0 
        is_late_930 = 0
    elif first_scan > limit_1000:
        is_late_1400 = 0
        is_late_1000 = 1
        is_late_930 = 0
    elif first_scan > limit_930:
        is_late_1400 = 0
        is_late_1000 = 0
        is_late_930 = 1

    is_saturday = str(row['day_str']).startswith('Sa')
    
    if is_late_1400:
        no_lunch = 0
    else:
        no_lunch = 1 if (len(times) < 4 and not is_saturday) else 0

    target = 4.0 if is_saturday else 8.0
    is_under = 1 if row['hours_worked'] > 0 and row['hours_worked'] < target else 0

    # --- HALF DAY LOGIC (REVISED) ---
    # Logic:
    # 1. NOT SATURDAY.
    # 2. Arrived >= 13:00 (Late Entry / Afternoon Only).
    #    OR
    # 3. Left <= 14:00 (Early Exit / Morning Only) AND Hours < 7 (To exclude continuous 7am-2pm shifts).
    
    if row['is_day_worked'] and not is_saturday and len(times) >= 2:
        try:
            t_entry = datetime.strptime(times[0], '%H:%M')
            t_exit = datetime.strptime(times[-1], '%H:%M')
            if t_exit < t_entry: t_exit += timedelta(days=1)
            
            # --- Condition A: Afternoon Only (Entered after 13:00) ---
            # Using 13:00 ensures we catch the 14:00 people too.
            cond_afternoon = (t_entry >= limit_1300)
            
            # --- Condition B: Morning Only (Left before 14:00) ---
            # Added hours check < 7 to ensure it's not a full continuous day
            cond_morning = (t_exit <= limit_1400) and (row['hours_worked'] < 7.0)
            
            if cond_afternoon or cond_morning:
                is_half_day = 1
                
        except:
            pass 

    return is_late_930, is_late_1000, is_late_1400, no_lunch, is_under, is_half_day

def calculate_business_days_in_range(start_date, end_date):
    current = start_date
    business_days = 0
    while current <= end_date:
        wd = current.weekday()
        if wd != 6:
            business_days += 1
        current += timedelta(days=1)
    return business_days

def minutes_to_hhmm(mins):
    if pd.isna(mins) or mins == 0:
        return ""
    hours = int(mins // 60)
    minutes = int(round(mins % 60))
    if minutes == 60:
        hours += 1
        minutes = 0
    return f"{hours:02}:{minutes:02}"

def decimal_hours_to_hhmm(decimal_hours):
    if pd.isna(decimal_hours) or decimal_hours == 0:
        return "00:00"
    
    is_negative = decimal_hours < 0
    minutes_total = abs(decimal_hours) * 60
    hours = int(minutes_total // 60)
    minutes = int(round(minutes_total % 60))
    
    if minutes == 60:
        hours += 1
        minutes = 0
        
    time_str = f"{hours:02}:{minutes:02}"
    return f"-{time_str}" if is_negative else time_str

def process_monthly_analysis(input_dir, output_dir):
    """
    Traite les fichiers dans input_dir et sauvegarde l'analyse mensuelle dans output_dir.
    Retourne le chemin du fichier généré ou None.
    """
    if not os.path.exists(input_dir):
        print(f"Dossier non trouvé : {input_dir}")
        return None

    # S'assurer que le dossier de sortie existe
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_data = []
    print("Reading files...")
    for file in os.listdir(input_dir):
        if file.lower().endswith(('.xls', '.xlsx')) and not file.startswith("Daily_Analysis") and not file.startswith("Monthly") and not file.startswith("Master") and not file.startswith("~$"):
            print(f"Processing: {file}")
            all_data.extend(extract_data(os.path.join(input_dir, file)))

    if not all_data:
        print("No data found.")
        return None

    df = pd.DataFrame(all_data)

    # --- DÉTECTION CHRONOLOGIQUE AMÉLIORÉE ---
    if 'day_numeric' in df.columns and not df.empty:
        # 1. Récupérer les infos de base
        month_num = df['month_num'].iloc[0] if 'month_num' in df.columns else '01'
        year_num = df['year_num'].iloc[0] if 'year_num' in df.columns else '2026'
        
        # 2. Identifier la séquence chronologique réelle
        unique_days_in_order = []
        seen = set()
        for d in df['day_numeric']:
            if d not in seen:
                unique_days_in_order.append(d)
                seen.add(d)

        real_start_day = unique_days_in_order[0]
        real_end_day = unique_days_in_order[-1]
        
        # Détecter s'il y a une transition de mois (ex: 25, 26... 31, 1, 2)
        has_transition = False
        pivot_index = -1
        for i in range(len(unique_days_in_order) - 1):
            if unique_days_in_order[i] > unique_days_in_order[i+1]:
                has_transition = True
                pivot_index = i
                break
        
        print(f"\n--- ANALYSE DE LA PÉRIODE ---")
        print(f"Séquence détectée : {unique_days_in_order}")
        
        # 3. Définir le jour cible (le dernier jour chronologique)
        target_report_day = real_end_day
        
        # 4. Vérifier si le dernier jour est complet (Scan count)
        last_day_records = df[df['day_numeric'] == target_report_day]
        total_last_day = len(last_day_records)
        # On considère un jour incomplet si + de 50% des gens n'ont qu'un seul pointage (ou 0)
        incomplete_count = len(last_day_records[last_day_records['scan_count'] <= 1])
        
        if total_last_day > 0 and (incomplete_count / total_last_day) > 0.5:
            print(f"DÉCISION : Le jour {target_report_day} est incomplet (en cours).")
            # Supprimer le jour incomplet du DataFrame pour l'analyse
            df = df[df['day_numeric'] != target_report_day].copy()
            # Le nouveau jour cible devient le précédent dans la liste ordonnée
            if len(unique_days_in_order) > 1:
                target_report_day = unique_days_in_order[-2]
                real_end_day = target_report_day
            print(f"Nouveau jour cible : {target_report_day}")
        else:
            print(f"DÉCISION : Le jour {target_report_day} est complet.")

        # 5. Calcul du nom du mois pour le header
        month_names = {
            '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
            '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
            '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
        }
        month_name = month_names.get(month_num, f'Mois {month_num}')
        
        # 6. Créer les dates complètes pour le calcul des jours ouvrés
        if 'full_date' in df.columns and not df['full_date'].isnull().all():
            # Utiliser les dates réelles si disponibles
            final_min_date = df['full_date'].min()
            final_max_date = df['full_date'].max()
        else:
            # Recréer les dates à partir des informations extraites
            final_min_date = datetime(int(year_num), int(month_num), real_start_day)
            final_max_date = datetime(int(year_num), int(month_num), real_end_day)
            
            # Gérer les périodes multi-mois
            if has_transition:
                # Si transition, le dernier mois est probablement le mois suivant
                if month_num == '12':
                    next_month_num = '01'
                    next_year_num = str(int(year_num) + 1)
                else:
                    next_month_num = f"{int(month_num) + 1:02d}"
                    next_year_num = year_num
                final_max_date = datetime(int(next_year_num), int(next_month_num), real_end_day)
        
        print(f"\n--- PLAGE DE JOURS DÉTECTÉE ---")
        print(f"Premier jour trouvé : {real_start_day}")
        print(f"Dernier jour trouvé : {real_end_day}")
        
        # Calculer correctement le total de jours pour les périodes multi-mois
        if has_transition:
            # Période multi-mois : jours du premier mois + jours du deuxième mois
            first_month_days = unique_days_in_order[:pivot_index + 1]
            second_month_days = unique_days_in_order[pivot_index + 1:]
            total_days = len(first_month_days) + len(second_month_days)
            print(f"Période multi-mois détectée : {len(first_month_days)} jours + {len(second_month_days)} jours")
        else:
            # Période simple
            total_days = len(unique_days_in_order)
        
        print(f"Total jours analysés : {total_days}")
        print(f"Final Analysis Period: {final_min_date.strftime('%d/%m/%Y')} to {final_max_date.strftime('%d/%m/%Y')}")
        global_expected_days = calculate_business_days_in_range(final_min_date, final_max_date)
        print(f"Theoretical Business Days (Mon-Sat) in period: {global_expected_days}")
        
        # Créer un nom de fichier dynamique basé sur la période analysée
        dynamic_filename = f"Monthly_Global_Analysis_{real_start_day:02d}-{month_num}-{year_num}_A_{real_end_day:02d}-{month_num}-{year_num}.xlsx"
        output_path = os.path.join(output_dir, dynamic_filename)
        header_text = f"Analyse Mensuelle - Période : {real_start_day} au {real_end_day} {month_name} {year_num}"

    else:
        print("Could not detect valid dates. Exiting.")
        return None

    if EXCLUDED_EMPLOYEES:
        print(f"\nFiltering out: {EXCLUDED_EMPLOYEES}")
        excluded_clean = [clean_name_string(name) for name in EXCLUDED_EMPLOYEES]
        df = df[~df['name'].isin(excluded_clean)]

    if df.empty:
        print("All data filtered out.")
        return None

    print("Analyzing metrics...")
    metrics = df.apply(analyze_record, axis=1)
    
    df['ENTRY > 9H30'] = [x[0] for x in metrics]
    df['ENTRY > 10H'] = [x[1] for x in metrics]
    df['ENTRY > 14H'] = [x[2] for x in metrics] 
    df['NO LUNCH'] = [x[3] for x in metrics]
    df['UNDER 8H'] = [x[4] for x in metrics]
    df['IS HALF DAY'] = [x[5] for x in metrics]

    report = df.groupby('name').agg({
        'is_day_worked': 'sum',
        'is_leave': 'sum',
        'is_holiday': 'sum',
        'daily_target_for_worked_day': 'sum', 
        'ENTRY > 10H': 'sum',
        'ENTRY > 14H': 'sum', 
        'ENTRY > 9H30': 'sum',
        'NO LUNCH': 'sum',
        'UNDER 8H': 'sum',
        'IS HALF DAY': 'sum',
        'hours_worked': 'sum',
        'daily_lunch_minutes': 'sum',
        'has_lunch_break': 'sum'
    }).reset_index()

    report.rename(columns={
        'name': 'Employee name',
        'is_day_worked': 'days worked',
        'daily_target_for_worked_day': 'TOTAL HOURS NEEDED', 
        'hours_worked': 'TOTAL HOURS WORKED',
        'IS HALF DAY': 'HALF DAYS' # Single numeric column
    }, inplace=True)

    report['real working days'] = global_expected_days - report['is_leave'] - report['is_holiday']
    report['ABSENCE'] = report['real working days'] - report['days worked']
    report['ABSENCE'] = report['ABSENCE'].apply(lambda x: max(0, x))
    
    report['avg_lunch_raw'] = report.apply(
        lambda x: x['daily_lunch_minutes'] / x['has_lunch_break'] if x['has_lunch_break'] > 0 else (x['daily_lunch_minutes'] if x['daily_lunch_minutes'] > 0 else 0), axis=1
    )
    report['AVG LUNCH TIME'] = report['avg_lunch_raw'].apply(minutes_to_hhmm)

    report['balance_raw'] = report['TOTAL HOURS WORKED'] - report['TOTAL HOURS NEEDED']
    report['Balance of hours worked'] = report['balance_raw'].apply(decimal_hours_to_hhmm)

    # --- EXPORT ---
    final_cols = [
        'Employee name', 
        'real working days', 
        'days worked',
        'ABSENCE', 
        'HALF DAYS', 
        'UNDER 8H', 
        'NO LUNCH', 
        'AVG LUNCH TIME',
        'ENTRY > 14H', 
        'ENTRY > 10H', 
        'ENTRY > 9H30', 
        'TOTAL HOURS NEEDED', 
        'TOTAL HOURS WORKED', 
        'Balance of hours worked'
    ]
    
    final_df = report[final_cols]

    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            # Ajouter l'en-tête sur la première ligne
            final_df.to_excel(writer, sheet_name='Monthly Summary', index=False, startrow=2, header=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Monthly Summary']
            
            # Format pour l'en-tête de période
            header_title = workbook.add_format({
                'bold': True, 'align': 'center', 'valign': 'vcenter',
                'font_size': 14, 'font_color': '#2F5597', 'border': 1
            })
            
            # Écrire l'en-tête de période sur la première ligne (fusionnée)
            if len(final_df.columns) > 1:
                worksheet.merge_range(0, 0, 0, len(final_df.columns) - 1, header_text, header_title)
            else:
                worksheet.write(0, 0, header_text, header_title)
            
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#4472C4', 'font_color': 'white', 'border': 1
            })
            
            header_red = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#C00000', 'font_color': 'white', 'border': 1
            })

            header_orange = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#ED7D31', 'font_color': 'white', 'border': 1
            })

            body_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
            text_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'})
            
            # Write Headers
            for col_num, value in enumerate(final_df.columns.values):
                if "14H" in value:
                     worksheet.write(1, col_num, value, header_red)
                elif "HALF DAYS" in value:
                    worksheet.write(1, col_num, value, header_orange)
                else:
                    worksheet.write(1, col_num, value, header_format)
            
            # Write Data
            for i, col in enumerate(final_df.columns):
                if col == 'Employee name':
                    cell_fmt = text_format
                    width = 20  # Reduced from 25
                elif col in ['real working days', 'days worked', 'ABSENCE', 'HALF DAYS', 'UNDER 8H', 'NO LUNCH', 'ENTRY > 14H', 'ENTRY > 10H', 'ENTRY > 9H30']:
                    cell_fmt = body_format
                    width = 10  # Count columns - narrower
                elif col in ['AVG LUNCH TIME']:
                    cell_fmt = body_format
                    width = 12  # Time column - medium width
                elif col in ['TOTAL HOURS NEEDED', 'TOTAL HOURS WORKED', 'Balance of hours worked']:
                    cell_fmt = body_format
                    width = 14  # Hour columns - slightly wider
                else:
                    cell_fmt = body_format
                    width = 12  # Default width
                
                worksheet.set_column(i, i, width)
                
                for row_idx, value in enumerate(final_df[col]):
                    if pd.isna(value): value = ""
                    # Show zeros for count columns, empty strings for time columns
                    if col in ['real working days', 'days worked', 'ABSENCE', 'HALF DAYS', 'UNDER 8H', 'NO LUNCH', 'ENTRY > 14H', 'ENTRY > 10H', 'ENTRY > 9H30']:
                        if value == 0: value = 0  # Keep zeros for count columns
                    elif col in ['AVG LUNCH TIME', 'Balance of hours worked', 'TOTAL HOURS WORKED']:
                        if value == 0 or value == "00:00": value = ""  # Empty string for time columns
                    worksheet.write(row_idx + 2, i, value, cell_fmt)

        print(f"\nSUCCESS! Monthly report generated: {output_path}")
        return output_path

    except Exception as e:
        print(f"Error saving file: {e}")
        return None

def main():
    if not os.path.exists(FOLDER_PATH):
        print("Folder not found.")
        return

    output = process_monthly_analysis(FOLDER_PATH, FOLDER_PATH)
    if output:
        print(f"Report generated: {output}")

if __name__ == "__main__":
    main()